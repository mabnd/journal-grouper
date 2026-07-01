"""
excel_grouper.py
================
Excel entry point for the journal-entry grouping algorithm. Each worksheet
in the input workbook is treated as one independent journal. The expected
input column names are defined in config.py (INPUT_COLUMNS).

All algorithm logic lives in journal_grouper_core.py — this file only
handles reading workbook sheets into row dicts and writing the result back
out as a new workbook (one output sheet per input sheet). See algorithm.md for the full step-by-step specification, and
csv_grouper.py for the equivalent single-journal CSV entry point (same
core logic, different file format).

Usage:
    python excel_grouper.py input.xlsx [clients.csv]

The output is written next to the input as <input>_processed.xlsx. If a
clients list is given (and exists), an extra sheet is appended listing
partner values not found in it, per source sheet.
"""

import os
import sys
import datetime

import openpyxl
from openpyxl.utils import get_column_letter

import journal_grouper_core as core
import config
from csv_grouper import derive_path, load_clients

# Column-width floor for autosize_columns(), in Excel's character units —
# no ceiling, so a column always fits its longest cell in full.
MIN_COLUMN_WIDTH = 8
COLUMN_WIDTH_PADDING = 2


def cell_to_text(value):
    """Converts a raw Excel cell value into the same kind of stripped text
    a CSV would have given us — handles numbers stored as numbers (not
    text) and dates stored as native date/datetime objects."""
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, datetime.date):
        return value.strftime(config.DATE_FORMAT)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value == int(value) else str(value)
    return str(value).strip()


def read_sheet(ws, sheet_name: str) -> tuple[list[dict], list[str]]:
    rows_iter = ws.iter_rows(values_only=True)
    header_raw = next(rows_iter, None)
    if header_raw is None:
        return [], []

    # Sheets are often padded with empty trailing columns (e.g. out to Z)
    # left over from formatting — trim those off rather than carrying a
    # pile of blank-named columns into the output.
    last_real = max(
        (i for i, h in enumerate(header_raw) if h is not None), default=-1
    )
    header_raw = header_raw[:last_real + 1]
    ncols = len(header_raw)
    fieldnames = core.normalize_fieldnames([cell_to_text(h) for h in header_raw])

    # If the sheet has no Journal column, the sheet name *is* the journal —
    # add the column ourselves rather than treating every row as missing it.
    has_journal_col = core.COL_JOURNAL in fieldnames
    if not has_journal_col:
        fieldnames = [core.COL_JOURNAL] + fieldnames

    rows = []
    idx = 0
    for raw_row in rows_iter:
        raw_row = raw_row[:ncols]
        if all(v is None for v in raw_row):
            continue
        row = {}
        if not has_journal_col:
            row[core.COL_JOURNAL] = sheet_name
        data_fields = fieldnames[1:] if not has_journal_col else fieldnames
        for col, value in zip(data_fields, raw_row):
            if col == core.COL_DEBIT or col == core.COL_CREDIT:
                row[col] = value
            else:
                row[col] = cell_to_text(value)
        row[core.COL_DATE] = core.normalize_date(row.get(core.COL_DATE, ""))
        row["_idx"]        = idx
        row["_debit_val"]  = core.parse_amount(row.get(core.COL_DEBIT))
        row["_credit_val"] = core.parse_amount(row.get(core.COL_CREDIT))
        rows.append(row)
        idx += 1
    return rows, fieldnames


def autosize_columns(ws):
    """Sets each column's width to fit its longest cell, within sane
    bounds — output columns no longer line up 1:1 with the source's own
    (renamed/reordered/merged), so mirroring the source's literal widths
    isn't meaningful anymore; this fits the content actually written."""
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = max(widths.get(cell.column, 0), len(str(cell.value)))
    for col, length in widths.items():
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = max(length + COLUMN_WIDTH_PADDING, MIN_COLUMN_WIDTH)


def write_sheet(ws_out, confirmed_groups, flagged_rows, fieldnames, resolution_meta):
    ncols = None
    output_fields = None
    for event, payload in core.generate_output_rows(
        confirmed_groups, flagged_rows, fieldnames, resolution_meta
    ):
        if event == "fields":
            output_fields = payload
            ncols = len(output_fields)
            ws_out.append(output_fields)
        elif event == "row":
            ws_out.append([payload.get(k, "") for k in output_fields])
        elif event == "blank":
            ws_out.append([None] * ncols)
        elif event == "header":
            ws_out.append([payload] + [None] * (ncols - 1))


def write_missing_clients_sheet(wb_out, issues_by_sheet: dict) -> None:
    ws = wb_out.create_sheet(title=config.MISSING_CLIENT_SHEET_NAME)
    ws.append([
        config.MISSING_CLIENT_COL_SHEET,
        config.MISSING_CLIENT_COL_PARTNER,
        config.MISSING_CLIENT_COL_OCCURRENCES,
        config.MISSING_CLIENT_COL_STATUS,
        config.MISSING_CLIENT_COL_SUGGESTION,
    ])
    for sheet_name, issues in issues_by_sheet.items():
        for name, (count, status, best_match) in sorted(issues.items()):
            status_text = config.MISSING_CLIENT_STATUS_TYPO if status == core.REASON_PARTNER_TYPO else config.MISSING_CLIENT_STATUS_UNKNOWN
            ws.append([sheet_name, name, count, status_text, best_match])
    autosize_columns(ws)


def run(input_path, clients_path=None):
    print(f"Reading: {input_path}")
    wb_in = openpyxl.load_workbook(input_path, data_only=True, read_only=True)

    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    known_names = None
    if clients_path:
        if not os.path.exists(clients_path):
            print(f"Clients list not found, skipping partner check: {clients_path}")
        else:
            known_names = load_clients(clients_path)

    total_confirmed   = 0
    total_flagged     = 0
    any_errors        = False
    issues_by_sheet: dict[str, dict] = {}
    skipped_sheets:  dict[str, list] = {}

    for sheet_name in wb_in.sheetnames:
        ws = wb_in[sheet_name]
        rows, fieldnames = read_sheet(ws, sheet_name)

        print(f"\n--- Sheet: {sheet_name} ---")
        if not rows:
            print(f"[{sheet_name}] Empty sheet, skipping")
            wb_out.create_sheet(title=sheet_name)
            continue

        missing = core.validate_required_columns(fieldnames)
        if missing:
            print(f"[{sheet_name}] SKIPPED — missing required column(s): {', '.join(missing)}")
            ws_out = wb_out.create_sheet(title=sheet_name)
            ws_out.append([f"SKIPPED — missing required column(s): {', '.join(missing)}"])
            skipped_sheets[sheet_name] = missing
            continue

        confirmed, flagged, resolution_meta, errors = core.process_and_report(
            rows, label=sheet_name
        )

        total_confirmed += len(confirmed)
        total_flagged   += len(flagged)
        any_errors       = any_errors or bool(errors)

        ws_out = wb_out.create_sheet(title=sheet_name)
        write_sheet(ws_out, confirmed, flagged, fieldnames, resolution_meta)
        autosize_columns(ws_out)

        if known_names is not None:
            issues = core.find_partner_issues(rows, known_names)
            if issues:
                issues_by_sheet[sheet_name] = issues

    print(f"\n=== Summary across {len(wb_in.sheetnames)} sheet(s) ===")
    print(f"  Total confirmed entries : {total_confirmed}")
    print(f"  Total flagged rows      : {total_flagged}")
    print(f"  Verification           : {'FAILED — see above' if any_errors else 'all sheets passed'}")
    if skipped_sheets:
        print(f"  Skipped sheets          : {len(skipped_sheets)} (missing required columns)")
        for sheet_name, missing in skipped_sheets.items():
            print(f"    - {sheet_name}: missing {', '.join(missing)}")

    if known_names is not None:
        if issues_by_sheet:
            write_missing_clients_sheet(wb_out, issues_by_sheet)
            total_issues = sum(len(u) for u in issues_by_sheet.values())
            total_typos = sum(
                1 for u in issues_by_sheet.values() for _, status, _ in u.values()
                if status == core.REASON_PARTNER_TYPO
            )
            print(
                f"  Partner issues          : {total_issues} "
                f"({total_typos} possible typo(s), {total_issues - total_typos} unknown) "
                f"— see '{config.MISSING_CLIENT_SHEET_NAME}' sheet"
            )
        else:
            print("  Partner issues          : none, all partners found in clients list")

    output_path = derive_path(input_path, "_processed")
    wb_out.save(output_path)
    print(f"\nOutput written to: {output_path}")


if __name__ == "__main__":
    if len(sys.argv) not in (2, 3):
        print("Usage: python excel_grouper.py <input.xlsx> [clients.csv]")
        sys.exit(1)
    run(sys.argv[1], sys.argv[2] if len(sys.argv) == 3 else None)
